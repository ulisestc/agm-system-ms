import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT


COLOR_HEADER    = "1A3C5E" 
COLOR_SUBHEADER = "2E6DA4"
COLOR_FILA_PAR  = "EAF2FB"
COLOR_TEXTO     = "FFFFFF"

def _borde_fino():
    lado = Side(style="thin", color="AAAAAA")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


# Excel - Calificaciones

def generar_excel_calificaciones(materia_id: str, datos: dict = None) -> tuple[bytes, str]:
    """
    Genera un archivo Excel con el concentrado de calificaciones de una materia.
    `datos` es un dict con la estructura:
        {
          "materia_nombre": str,
          "materia_nrc": str,
          "periodo": str,
          "docente": str,
          "alumnos": [
            { "matricula": str, "nombre": str, "calificacion_final": float, "promedio_real": float }
          ]
        }
    Si datos es None se genera con datos de demostración (útil para pruebas).
    """
    if datos is None:
        datos = _datos_demo_calificaciones(materia_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Concentrado Calificaciones"

    # Encabezado institucional
    ws.merge_cells("A1:F1")
    ws["A1"] = "BENEMÉRITA UNIVERSIDAD AUTÓNOMA DE PUEBLA"
    ws["A1"].font       = Font(bold=True, size=13, color=COLOR_TEXTO)
    ws["A1"].fill       = PatternFill("solid", fgColor=COLOR_HEADER)
    ws["A1"].alignment  = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:F2")
    ws["A2"] = "Facultad de Ciencias de la Computación — AGM Sistema de Calificaciones"
    ws["A2"].font      = Font(bold=True, size=10, color=COLOR_TEXTO)
    ws["A2"].fill      = PatternFill("solid", fgColor=COLOR_SUBHEADER)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Datos de la materia
    info_filas = [
        ("Materia:",  datos.get("materia_nombre", "—")),
        ("NRC:",      datos.get("materia_nrc", "—")),
        ("Periodo:",  datos.get("periodo", "—")),
        ("Docente:",  datos.get("docente", "—")),
        ("Fecha de generación:", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    fila_actual = 3
    for etiqueta, valor in info_filas:
        ws[f"A{fila_actual}"] = etiqueta
        ws[f"A{fila_actual}"].font = Font(bold=True)
        ws[f"B{fila_actual}"] = valor
        ws.merge_cells(f"B{fila_actual}:F{fila_actual}")
        fila_actual += 1

    fila_actual += 1 

    # Encabezado de la tabla
    encabezados = ["#", "Matrícula", "Nombre del Alumno", "Promedio Real", "Calificación Final", "Estado"]
    for col_idx, titulo in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila_actual, column=col_idx, value=titulo)
        celda.font      = Font(bold=True, color=COLOR_TEXTO)
        celda.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border    = _borde_fino()
    ws.row_dimensions[fila_actual].height = 18
    fila_actual += 1

    # Filas de alumnos
    for idx, alumno in enumerate(datos.get("alumnos", []), start=1):
        cal_final = alumno.get("calificacion_final", 0)
        estado    = "APROBADO" if cal_final >= 6 else "REPROBADO"
        color_fila = COLOR_FILA_PAR if idx % 2 == 0 else "FFFFFF"

        fila_datos = [
            idx,
            alumno.get("matricula", "—"),
            alumno.get("nombre", "—"),
            round(alumno.get("promedio_real", 0), 2),
            cal_final,
            estado,
        ]
        for col_idx, valor in enumerate(fila_datos, start=1):
            celda = ws.cell(row=fila_actual, column=col_idx, value=valor)
            celda.fill      = PatternFill("solid", fgColor=color_fila)
            celda.border    = _borde_fino()
            celda.alignment = Alignment(horizontal="center" if col_idx != 3 else "left")
            if estado == "REPROBADO" and col_idx == 6:
                celda.font = Font(color="C0392B", bold=True)
            elif estado == "APROBADO" and col_idx == 6:
                celda.font = Font(color="1E8449", bold=True)
        fila_actual += 1

    # Anchos de columna
    anchos = [5, 14, 36, 16, 20, 14]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nrc      = datos.get("materia_nrc", materia_id)
    filename = f"calificaciones_{nrc}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return buffer.read(), filename


# Excel - Asistencias

def generar_excel_asistencias(materia_id: str, datos: dict = None) -> tuple[bytes, str]:
    """
    Genera Excel con el concentrado de asistencias.
    `datos` debe tener la misma cabecera de materia más:
        "sesiones": [ { "fecha": str, "alumnos": [ { "nombre": str, "matricula": str, "estado": str } ] } ]
    """
    if datos is None:
        datos = _datos_demo_asistencias(materia_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Concentrado Asistencias"

    ws.merge_cells("A1:E1")
    ws["A1"] = "BENEMÉRITA UNIVERSIDAD AUTÓNOMA DE PUEBLA — Concentrado de Asistencias"
    ws["A1"].font      = Font(bold=True, size=12, color=COLOR_TEXTO)
    ws["A1"].fill      = PatternFill("solid", fgColor=COLOR_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 20

    fila_actual = 2
    for etiqueta, valor in [
        ("Materia:", datos.get("materia_nombre", "—")),
        ("NRC:", datos.get("materia_nrc", "—")),
        ("Periodo:", datos.get("periodo", "—")),
    ]:
        ws[f"A{fila_actual}"] = etiqueta
        ws[f"A{fila_actual}"].font = Font(bold=True)
        ws[f"B{fila_actual}"] = valor
        ws.merge_cells(f"B{fila_actual}:E{fila_actual}")
        fila_actual += 1

    fila_actual += 1

    for sesion in datos.get("sesiones", []):
        ws.merge_cells(f"A{fila_actual}:E{fila_actual}")
        ws[f"A{fila_actual}"] = f"Sesión del {sesion.get('fecha', '—')}"
        ws[f"A{fila_actual}"].font = Font(bold=True, color=COLOR_TEXTO)
        ws[f"A{fila_actual}"].fill = PatternFill("solid", fgColor=COLOR_SUBHEADER)
        ws[f"A{fila_actual}"].alignment = Alignment(horizontal="center")
        fila_actual += 1

        encabezados = ["#", "Matrícula", "Nombre", "Estado", "Hora registro"]
        for col_idx, titulo in enumerate(encabezados, start=1):
            celda = ws.cell(row=fila_actual, column=col_idx, value=titulo)
            celda.font   = Font(bold=True)
            celda.border = _borde_fino()
            celda.alignment = Alignment(horizontal="center")
        fila_actual += 1

        for idx, alumno in enumerate(sesion.get("alumnos", []), start=1):
            estado = alumno.get("estado", "Presente")
            color_estado = {"Presente": "1E8449", "Retardo": "D68910", "Falta": "C0392B"}.get(estado, "000000")
            fila_datos = [idx, alumno.get("matricula", ""), alumno.get("nombre", ""), estado, alumno.get("hora", "")]
            for col_idx, valor in enumerate(fila_datos, start=1):
                celda = ws.cell(row=fila_actual, column=col_idx, value=valor)
                celda.border = _borde_fino()
                celda.alignment = Alignment(horizontal="center" if col_idx != 3 else "left")
                if col_idx == 4:
                    celda.font = Font(color=color_estado, bold=True)
            fila_actual += 1

        fila_actual += 1

    for i, ancho in enumerate([5, 14, 34, 14, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nrc      = datos.get("materia_nrc", materia_id)
    filename = f"asistencias_{nrc}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return buffer.read(), filename


# PDF - Calificaciones

def generar_pdf_calificaciones(materia_id: str, datos: dict = None) -> tuple[bytes, str]:
    if datos is None:
        datos = _datos_demo_calificaciones(materia_id)

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=letter,
                               topMargin=0.6*inch, bottomMargin=0.6*inch,
                               leftMargin=0.7*inch, rightMargin=0.7*inch)
    styles  = getSampleStyleSheet()
    historia = []

    # Título
    estilo_titulo = ParagraphStyle("titulo", parent=styles["Heading1"],
                                   alignment=TA_CENTER, fontSize=14,
                                   textColor=colors.HexColor(f"#{COLOR_HEADER}"))
    historia.append(Paragraph("BENEMÉRITA UNIVERSIDAD AUTÓNOMA DE PUEBLA", estilo_titulo))
    historia.append(Paragraph("Facultad de Ciencias de la Computación — AGM", styles["Normal"]))
    historia.append(Spacer(1, 0.1*inch))

    # Info de la materia
    estilo_info = ParagraphStyle("info", parent=styles["Normal"], fontSize=10)
    for etiqueta, valor in [
        ("Materia", datos.get("materia_nombre", "—")),
        ("NRC", datos.get("materia_nrc", "—")),
        ("Periodo", datos.get("periodo", "—")),
        ("Docente", datos.get("docente", "—")),
        ("Fecha", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]:
        historia.append(Paragraph(f"<b>{etiqueta}:</b> {valor}", estilo_info))
    historia.append(Spacer(1, 0.2*inch))

    # Tabla de calificaciones
    tabla_datos = [["#", "Matrícula", "Nombre del Alumno", "Prom. Real", "Cal. Final", "Estado"]]
    for idx, alumno in enumerate(datos.get("alumnos", []), start=1):
        cal  = alumno.get("calificacion_final", 0)
        tabla_datos.append([
            str(idx),
            alumno.get("matricula", "—"),
            alumno.get("nombre", "—"),
            str(round(alumno.get("promedio_real", 0), 2)),
            str(cal),
            "APROBADO" if cal >= 6 else "REPROBADO",
        ])

    tabla = Table(tabla_datos, colWidths=[0.4*inch, 1.0*inch, 2.5*inch, 0.9*inch, 0.9*inch, 1.0*inch])
    tabla.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor(f"#{COLOR_HEADER}")),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  9),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("ALIGN",         (2, 1), (2, -1),  "LEFT"),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{COLOR_FILA_PAR}")]),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#AAAAAA")),
        ("FONTSIZE",      (0, 1), (-1, -1), 8),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    historia.append(tabla)

    doc.build(historia)
    buffer.seek(0)

    nrc      = datos.get("materia_nrc", materia_id)
    filename = f"calificaciones_{nrc}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return buffer.read(), filename


# PDF - Asistencias

def generar_pdf_asistencias(materia_id: str, datos: dict = None) -> tuple[bytes, str]:
    if datos is None:
        datos = _datos_demo_asistencias(materia_id)

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                               topMargin=0.6*inch, bottomMargin=0.6*inch,
                               leftMargin=0.7*inch, rightMargin=0.7*inch)
    styles   = getSampleStyleSheet()
    historia = []

    estilo_titulo = ParagraphStyle("titulo", parent=styles["Heading1"],
                                   alignment=TA_CENTER, fontSize=13,
                                   textColor=colors.HexColor(f"#{COLOR_HEADER}"))
    historia.append(Paragraph("Concentrado de Asistencias — AGM BUAP", estilo_titulo))

    for etiqueta, valor in [
        ("Materia", datos.get("materia_nombre", "—")),
        ("NRC", datos.get("materia_nrc", "—")),
        ("Periodo", datos.get("periodo", "—")),
    ]:
        historia.append(Paragraph(f"<b>{etiqueta}:</b> {valor}", styles["Normal"]))
    historia.append(Spacer(1, 0.15*inch))

    for sesion in datos.get("sesiones", []):
        historia.append(Paragraph(f"<b>Sesión: {sesion.get('fecha', '—')}</b>", styles["Heading3"]))
        tabla_datos = [["#", "Matrícula", "Nombre", "Estado", "Hora"]]
        for idx, alumno in enumerate(sesion.get("alumnos", []), start=1):
            tabla_datos.append([
                str(idx),
                alumno.get("matricula", ""),
                alumno.get("nombre", ""),
                alumno.get("estado", ""),
                alumno.get("hora", ""),
            ])

        tabla = Table(tabla_datos, colWidths=[0.4*inch, 1.1*inch, 3.0*inch, 1.0*inch, 1.0*inch])
        tabla.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor(f"#{COLOR_SUBHEADER}")),
            ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
            ("ALIGN",         (2, 1), (2, -1),  "LEFT"),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{COLOR_FILA_PAR}")]),
            ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#AAAAAA")),
            ("FONTSIZE",      (0, 0), (-1, -1), 8),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ]))
        historia.append(tabla)
        historia.append(Spacer(1, 0.15*inch))

    doc.build(historia)
    buffer.seek(0)

    nrc      = datos.get("materia_nrc", materia_id)
    filename = f"asistencias_{nrc}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return buffer.read(), filename


# Datos de demostración (pruebas sin integración)

def _datos_demo_calificaciones(materia_id: str) -> dict:
    return {
        "materia_nombre": f"Materia Demo [{materia_id}]",
        "materia_nrc": "12345",
        "periodo": "Primavera 2025",
        "docente": "Docente Demo",
        "alumnos": [
            {"matricula": "201900001", "nombre": "Ana García López",     "promedio_real": 9.3,  "calificacion_final": 9},
            {"matricula": "201900002", "nombre": "Luis Martínez Ruiz",   "promedio_real": 7.6,  "calificacion_final": 8},
            {"matricula": "201900003", "nombre": "María Pérez Castillo", "promedio_real": 5.4,  "calificacion_final": 5},
            {"matricula": "201900004", "nombre": "Carlos Sánchez Vega",  "promedio_real": 8.8,  "calificacion_final": 9},
        ]
    }

def _datos_demo_asistencias(materia_id: str) -> dict:
    return {
        "materia_nombre": f"Materia Demo [{materia_id}]",
        "materia_nrc": "12345",
        "periodo": "Primavera 2025",
        "sesiones": [
            {
                "fecha": "2025-04-10",
                "alumnos": [
                    {"matricula": "201900001", "nombre": "Ana García López",    "estado": "Presente", "hora": "09:02"},
                    {"matricula": "201900002", "nombre": "Luis Martínez Ruiz",  "estado": "Retardo",  "hora": "09:08"},
                    {"matricula": "201900003", "nombre": "María Pérez Castillo","estado": "Falta",    "hora": "—"},
                ]
            },
        ]
    }
