"""
Tests unitarios para ms-reportes: transformación de asistencias y generadores PDF/Excel.
Ejecutar con: pytest tests/test_reportes.py -v (desde el directorio ms-reportes)
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from unittest.mock import patch, MagicMock


# ── construir_sesiones_asistencia ─────────────────────────────────────────────

def _make_mock_rpc():
    """Crea un mock de RabbitMQRpcClient que no intenta conectarse."""
    mock = MagicMock()
    mock.return_value = mock
    return mock


def _build_construir_sesiones(asistencias_data: dict):
    """Importa construir_sesiones_asistencia con rpc_client mockeado."""
    mock_rpc_instance = MagicMock()

    def call_side_effect(queue_name, action, data, **kwargs):
        alumno_id = data.get("alumnoId")
        return {"success": True, "asistencias": asistencias_data.get(alumno_id, [])}

    mock_rpc_instance.call.side_effect = call_side_effect

    with patch.dict("sys.modules", {
        "rabbitmq_manager": MagicMock(
            RabbitMQRpcClient=MagicMock(return_value=mock_rpc_instance),
            RabbitMQManager=MagicMock(return_value=MagicMock()),
        )
    }):
        if "rabbitmq_client" in sys.modules:
            del sys.modules["rabbitmq_client"]
        import rabbitmq_client
        rabbitmq_client.rpc_client = mock_rpc_instance
        return rabbitmq_client.construir_sesiones_asistencia


def test_construir_sesiones_agrupa_por_fecha():
    """Las asistencias deben agruparse por fecha extraída de hora_registro."""
    alumnos = [
        {"id": "1", "matricula": "202300001", "nombre": "Ana García"},
        {"id": "2", "matricula": "202300002", "nombre": "Luis Torres"},
    ]
    asistencias_data = {
        "1": [
            {"estado": "Presente", "hora_registro": "2025-04-10 09:00:00"},
            {"estado": "Retardo",  "hora_registro": "2025-04-17 09:10:00"},
        ],
        "2": [
            {"estado": "Presente", "hora_registro": "2025-04-10 09:05:00"},
        ],
    }

    construir = _build_construir_sesiones(asistencias_data)
    result = construir(alumnos, materia_db_id="15")

    assert len(result) == 2, "Debe haber 2 fechas de sesión"
    fechas = {s["fecha"] for s in result}
    assert "2025-04-10" in fechas
    assert "2025-04-17" in fechas

    sesion_10 = next(s for s in result if s["fecha"] == "2025-04-10")
    assert len(sesion_10["alumnos"]) == 2
    matriculas = {a["matricula"] for a in sesion_10["alumnos"]}
    assert "202300001" in matriculas
    assert "202300002" in matriculas

    sesion_17 = next(s for s in result if s["fecha"] == "2025-04-17")
    assert len(sesion_17["alumnos"]) == 1
    assert sesion_17["alumnos"][0]["estado"] == "Retardo"


def test_construir_sesiones_usa_matricula_no_id():
    """El campo 'matricula' del reporte debe ser alumno['matricula'], no alumno['id']."""
    alumnos = [{"id": "99", "matricula": "202399999", "nombre": "Test"}]
    asistencias_data = {
        "99": [{"estado": "Presente", "hora_registro": "2025-05-01 10:00:00"}]
    }
    construir = _build_construir_sesiones(asistencias_data)
    result = construir(alumnos, materia_db_id="10")

    assert result[0]["alumnos"][0]["matricula"] == "202399999"


def test_construir_sesiones_sin_asistencias():
    """Si no hay registros, se devuelve lista vacía (no error)."""
    alumnos = [{"id": "1", "matricula": "202300001", "nombre": "Ana"}]
    construir = _build_construir_sesiones({"1": []})
    result = construir(alumnos, materia_db_id="5")
    assert result == []


def test_construir_sesiones_ordena_por_fecha():
    """Las sesiones deben estar ordenadas cronológicamente."""
    alumnos = [{"id": "1", "matricula": "202300001", "nombre": "Ana"}]
    asistencias_data = {
        "1": [
            {"estado": "Presente", "hora_registro": "2025-04-17 09:00:00"},
            {"estado": "Presente", "hora_registro": "2025-04-03 09:00:00"},
            {"estado": "Presente", "hora_registro": "2025-04-10 09:00:00"},
        ]
    }
    construir = _build_construir_sesiones(asistencias_data)
    result = construir(alumnos, materia_db_id="1")

    fechas = [s["fecha"] for s in result]
    assert fechas == sorted(fechas), "Las fechas deben estar en orden ascendente"


# ── Generadores PDF/Excel (smoke tests, sin RabbitMQ) ────────────────────────

def test_generar_excel_calificaciones_con_datos_reales():
    """El Excel de calificaciones se genera sin error con datos reales."""
    from generadores import generar_excel_calificaciones
    datos = {
        "materia_nombre": "Programación",
        "materia_nrc": "12345",
        "periodo": "2025-A",
        "docente": "Prof. Pérez",
        "alumnos": [
            {"matricula": "202300001", "nombre": "Ana García", "promedio_real": 8.5, "calificacion_final": 9},
            {"matricula": "202300002", "nombre": "Luis Torres", "promedio_real": 5.3, "calificacion_final": 5},
        ]
    }
    file_bytes, filename = generar_excel_calificaciones("12345", datos)
    assert len(file_bytes) > 0
    assert filename.endswith(".xlsx")
    assert "12345" in filename


def test_generar_excel_calificaciones_alumno_reprobado_aprobado():
    """El estado APROBADO/REPROBADO se determina por calificacion_final >= 6."""
    from generadores import generar_excel_calificaciones
    import openpyxl, io
    datos = {
        "materia_nombre": "Test",
        "materia_nrc": "99999",
        "periodo": "2025-A",
        "docente": "Doc",
        "alumnos": [
            {"matricula": "001", "nombre": "Aprobado",  "promedio_real": 7.0, "calificacion_final": 7},
            {"matricula": "002", "nombre": "Reprobado", "promedio_real": 4.0, "calificacion_final": 4},
        ]
    }
    file_bytes, _ = generar_excel_calificaciones("99999", datos)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    # Buscar filas de datos (después del encabezado)
    estados = [ws.cell(row=r, column=6).value for r in range(1, ws.max_row + 1)
               if ws.cell(row=r, column=6).value in ("APROBADO", "REPROBADO")]
    assert "APROBADO" in estados
    assert "REPROBADO" in estados


def test_generar_excel_asistencias_estructura_correcta():
    """El Excel de asistencias se genera con la estructura date-grouped."""
    from generadores import generar_excel_asistencias
    datos = {
        "materia_nombre": "Programación",
        "materia_nrc": "12345",
        "periodo": "2025-A",
        "sesiones": [
            {
                "fecha": "2025-04-10",
                "alumnos": [
                    {"matricula": "202300001", "nombre": "Ana", "estado": "Presente", "hora": "09:00"},
                    {"matricula": "202300002", "nombre": "Luis", "estado": "Falta", "hora": "—"},
                ]
            }
        ]
    }
    file_bytes, filename = generar_excel_asistencias("12345", datos)
    assert len(file_bytes) > 0
    assert filename.endswith(".xlsx")
